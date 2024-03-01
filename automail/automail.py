from datetime import datetime
from email.header import Header
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from os.path import basename
import smtplib
from email.mime.text import MIMEText

wb = load_workbook('companyemails.xlsx')

sheet1 = wb.active

coverLetterTemplate = """
<!DOCTYPE html>
<html>

<head>
    <title>Cover Letter</title>
    <style>
        body {{
            font-family: 'Times New Roman', serif;
            margin: 40px;
            font-size: 12pt;
        }}

        a {{text-decoration: none;
            color: black;
            }}

        .date {{ text-align: right;}}
           
        

        .content {{margin-top: 20px;}}
            
        

        .signature {{margin-top: 40px;
            text-align: left;}}
            

        .contact-info {{text-align: left;}}
            

        .contact-info p {{margin: 2px;}}
            
    </style>
</head>

<body>
    <div class="contact-info">

        <p>name</p>
        <p>cellphone </p> 
        <p><a href="mailto:email">email</a></p>
        <p>Personal website: </p>
        <p><a href="website">website</a></p>
        <p>{date}</p>
    </div>
    <div class="content">

        <p>{companyName}</p>
        <p>Dear Hiring Manager,</p>
        <p>I am writing to express my deep interest in the software development opportunities within {companyName}. I
            have carefully read your request and I think I am qualified for the job. I have a strong drive to learn and
            a great passion for programming. My journey through rigorous academic training and a passionate self-driven
            exploration of programming has uniquely prepared me for a challenging and rewarding role in {companyName}.
        </p>
       

    </div>
    </div>
    <div class="signature">
        <p>Sincerely,</p>
        <p>name</p>
    </div>
</body>

</html>
"""
count = 0
for row in sheet1.iter_rows():
    if count == 0:
        count += 1
        continue
    companyName = f'{row[0].value.title()}'
    companyEmail = f'{row[1].value}'
    techRequired = f'{row[2].value}'
    coverLetter = coverLetterTemplate.format(
        date = datetime.now().strftime('%Y-%m-%d'),
        companyName=companyName
    ) 
    count += 1
#pdf path
    attachment_file = 'CV.pdf'
        #send mail server
    smtp_obj = smtplib.SMTP_SSL('smtp.gmail.com', 465)

            #login the email
    smtp_obj.login('personal email','password')

            #the email content
     # create MIMEMultipart objects
    msg = MIMEMultipart()
            #to where the email
    msg['From'] = Header('header','utf-8')
            #object
    msg['Subject'] = Header('job position','utf-8')
    # set the cover letter in text content
    msg.attach(MIMEText(coverLetter, 'html', 'utf-8'))
   # read PDF and add to attachments
    with open(attachment_file, 'rb') as f:
        part = MIMEApplication(f.read(), _subtype="pdf")
        part.add_header('Content-Disposition', 'attachment', filename=basename(attachment_file))
        msg.attach(part)


    smtp_obj.sendmail('personal email',[companyEmail],msg.as_string())
    print(f'success to send mail：{companyName} email address: {companyEmail}')

# exit SMTP 
smtp_obj.quit()